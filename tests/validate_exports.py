from datetime import timedelta
from pathlib import Path
from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / 'qa-output'
BLANKS = ['H20', 'H21', 'H22', 'I20', 'I21', 'I22', 'J20', 'J21', 'J22']


def total_minutes(value):
    if isinstance(value, timedelta):
        return round(value.total_seconds() / 60)
    return int(value)


def check_workbook(path, sheet_count):
    wb = load_workbook(path, data_only=False)
    assert len(wb.sheetnames) == sheet_count
    for ws in wb.worksheets:
        assert ws.max_row == 22
        assert ws.max_column == 24
        assert all(ws[cell].value is None for cell in BLANKS)
        assert sum(ws.row_dimensions[row].height for row in range(1, 23)) == 425
        assert ws.page_setup.orientation == 'landscape'
        assert str(ws.page_setup.paperSize) == '8'
        assert str(ws.print_area).endswith("!$A$1:$X$22")
    first = wb.worksheets[0]
    assert total_minutes(first['K21'].value) == 784 * 60 + 37
    assert first['M21'].value == 1455
    assert total_minutes(first['Q21'].value) == 682 * 60 + 29
    assert total_minutes(first['S21'].value) == 92 * 60 + 9
    assert total_minutes(first['T21'].value) == 9 * 60 + 59
    for cell in ['K21', 'Q21', 'S21', 'T21']:
        assert first[cell].number_format == '[h]:mm'
    wb.close()


def check_pdf(path, page_count):
    reader = PdfReader(path)
    assert len(reader.pages) == page_count
    for page in reader.pages:
        width = float(page.mediabox.width)
        height = float(page.mediabox.height)
        assert abs(width - 1190.551) < 0.02
        assert abs(height - 841.89) < 0.02


check_workbook(OUT / 'LibrettoVolo_Empty_QA.xlsx', 1)
check_workbook(OUT / 'LibrettoVolo_QA.xlsx', 1)
check_workbook(OUT / 'LibrettoVolo_204_QA.xlsx', 13)
check_pdf(OUT / 'LibrettoVolo_Empty_QA.pdf', 1)
check_pdf(OUT / 'LibrettoVolo_QA.pdf', 1)
check_pdf(OUT / 'LibrettoVolo_204_QA.pdf', 13)
print('Export Excel e PDF verificati.')
